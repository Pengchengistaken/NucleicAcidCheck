import os
import sys
import openpyxl
import copy
from openpyxl.styles import Alignment, PatternFill, NamedStyle, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from paddleocr import PaddleOCR
import pandas as pd
import re
import time
import process_input
from pprint import pprint
import numpy as np
import fire

ocr = PaddleOCR(use_gpu=False, use_angle_cls=True, lang="ch")
info_dict = {}
validate_date_delta = 2
notice_date = '20220605'


def check(date, time_ocr):
    if get_date_delta(time_ocr, date) > validate_date_delta:
        return False
    return True


def match(reg, total):
    reg_re = re.search(reg, total, re.M | re.I)
    reg_ocr = ""
    if reg_re and len(reg_re.groups()) >= 1:
        reg_ocr = reg_re.group(1)
    return reg_ocr


def get_date_delta(day1, day2):
    time_array1 = time.strptime(day1, "%Y.%m.%d")
    timestamp_day1 = int(time.mktime(time_array1))
    time_array2 = time.strptime(day2, "%Y.%m.%d")
    timestamp_day2 = int(time.mktime(time_array2))
    return (timestamp_day2 - timestamp_day1) // 60 // 60 // 24


# 通过对比转换后的日期数字判断日期范围
def check_date(start_date, last_date, date_orc):
    print(start_date, last_date, date_orc)
    if format_date(start_date) <= format_date(date_orc) <= format_date(last_date):
        return True
    return False


# 将日期字符串转换成数字
def format_date(date):
    date = re.sub(r'\D', "", date)[:8]
    print('日期转换：', date)
    if len(date) < 8:
        return 20220101
    return int(date)


def save_to_file(df, file_name):
    output_file_name = "./output/check_" + file_name
    book = openpyxl.Workbook()
    sheet1 = book.active
    df.index += 1  # 序号从 1 开始
    for row in dataframe_to_rows(df, index=False):
        sheet1.append(row)

    # 样式_标题行样式
    style_title_row = NamedStyle(name='style_title_row',
                                 font=Font(b=True),  # 粗体
                                 fill=PatternFill(fill_type='solid',  # 指定填充的类型，支持的有：'solid'等。
                                                  start_color='cccccc',  # 指定填充的开始颜色
                                                  end_color='cccccc'  # 指定填充的结束颜色
                                                  ),
                                 alignment=Alignment(horizontal='center',  # 水平居中
                                                     vertical='center',  # 垂直居中
                                                     wrap_text=True,  # 自动换行
                                                     )
                                 )

    # 边框样式
    line_t = Side(style='thin', color='000000')  # 细边框
    line_m = Side(style='medium', color='000000')  # 粗边框
    border = Border(top=line_m, bottom=line_t, left=line_t, right=line_t)
    style_border = NamedStyle(name='style_border', border=border)
    # 设置填充颜色
    colors = ['ffc7ce', 'c6efce', 'ffeb9c']  # 红 绿 黄
    fill_red = PatternFill('solid', fgColor=colors[0])  # 设置填充颜色为 橙红
    fill_green = PatternFill('solid', fgColor=colors[1])  # 设置填充颜色为 绿色
    fill_yellow = PatternFill('solid', fgColor=colors[2])  # 设置填充颜色为 黄色

    # 冻结第一行
    sheet1.freeze_panes = 'A2'

    # 设置列宽度
    for i in range(1, sheet1.max_column + 1):
        sheet1.column_dimensions[get_column_letter(i)].width = 30

    # 设置姓名列宽度
    name_cols_list = ['A', 'C', 'F', 'I', 'L', 'O']
    if notice_date == '20220605':
        name_cols_list = ['A', 'C', 'E', 'G', 'I', 'K']
    for col in name_cols_list:
        sheet1.column_dimensions[col].width = 12

    # 按行进行设置
    for row in sheet1.iter_rows():
        for cell in row:
            # 设置边框
            cell.style = style_border
            # 自动换行
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment
            # 填充颜色
            for name_col in name_cols_list:
                if name_col in cell.coordinate:
                    if str(cell.value) != 'nan':
                        print("单元格的值是：" + cell.value)
                        cell.fill = fill_yellow
            if cell.value is not None:
                value = str(cell.value)
                if value.find("不及格") != -1 or value.find("注意") != -1:
                    cell.fill = fill_red
                elif value.find("及格") != -1:
                    cell.fill = fill_green
    # 设置标题样式
    for row in sheet1['A1:R1']:  # 设置标题行样式
        for cell in row:
            cell.style = style_title_row

    book.save(output_file_name)  # 保存


def deal_file(notice_date='20220605', file_path='503-20220605-20220605.xlsx'):
    # 定义输出的Excel表格的各个栏位
    if notice_date == '20220605':
        df = pd.DataFrame(columns=[
            "学生姓名",
            "学生核酸图片结果",
            # "学生行程码图片结果",
            "同住人1姓名",
            # "同住人1核酸图片结果",
            "同住人1行程码图片结果",
            "同住人2姓名",
            # "同住人2核酸图片结果",
            "同住人2行程码图片结果",
            "同住人3姓名",
            # "同住人3核酸图片结果",
            "同住人3行程码图片结果",
            "同住人4姓名",
            # "同住人4核酸图片结果",
            "同住人4行程码图片结果",
            "同住人5姓名",
            # "同住人5核酸图片结果",
            "同住人5行程码图片结果",
            "备注说明"
        ])
    else:
        df = pd.DataFrame(columns=[
            "学生姓名",
            "学生核酸图片结果",
            "同住人1姓名",
            "同住人1核酸图片结果",
            "同住人1行程码图片结果",
            "同住人2姓名",
            "同住人2核酸图片结果",
            "同住人2行程码图片结果",
            "同住人3姓名",
            "同住人3核酸图片结果",
            "同住人3行程码图片结果",
            "同住人4姓名",
            "同住人4核酸图片结果",
            "同住人4行程码图片结果",
            "同住人5姓名",
            "同住人5核酸图片结果",
            "同住人5行程码图片结果",
            "备注说明"
        ])

    # 获得学生及同住人信息
    if notice_date == '20220605':
        info_list = process_input.read_excel_info(file_path, img_col_index=[3, 5, 7, 9, 11, 13])
    else:
        info_list = process_input.read_excel_info(file_path, img_col_index=[3, 5, 6, 8, 9, 11, 12, 14, 15, 17, 18])
    # 打印列表
    pprint(info_list)

    # 提取学生及同住人信息
    for student_info in info_list:
        if notice_date == '20220605':
            student_name = student_info.get('学生姓名')
            student_result_image = student_info.get('学生6月5日当天检测的24小时核酸检测结果截图')
            # student_travel_image = student_info.get('学生5月15日当天的行程卡截图')
            student_relative1_name = student_info.get('同住人1的姓名')
            # student_relative1_result_image = student_info.get('同住人1的24小时核酸检测结果截图')
            student_relative1_travel_image = student_info.get('同住人1行程卡截图')
            student_relative2_name = student_info.get('同住人2的姓名')
            # student_relative2_result_image = student_info.get('同住人2的24小时核酸检测结果截图')
            student_relative2_travel_image = student_info.get('同住人2行程卡截图')
            student_relative3_name = student_info.get('同住人3的姓名')
            # student_relative3_result_image = student_info.get('同住人3的24小时核酸检测结果截图')
            student_relative3_travel_image = student_info.get('同住人3行程卡截图')
            student_relative4_name = student_info.get('同住人4的姓名')
            # student_relative4_result_image = student_info.get('同住人4的24小时核酸检测结果截图')
            student_relative4_travel_image = student_info.get('同住人4行程卡截图')
            student_relative5_name = student_info.get('同住人5的姓名')
            # student_relative5_result_image = student_info.get('同住人5的24小时核酸检测结果截图')
            student_relative5_travel_image = student_info.get('同住人5行程卡截图')
        else:
            student_name = student_info.get('学生姓名')
            student_result_image = student_info.get('学生的24小时核酸检测结果截图')
            student_relative1_name = student_info.get('同住人1的姓名')
            student_relative1_result_image = student_info.get('同住人1的24小时核酸检测结果截图')
            student_relative1_travel_image = student_info.get('同住人1行程码截图')
            student_relative2_name = student_info.get('同住人2的姓名')
            student_relative2_result_image = student_info.get('同住人2的24小时核酸检测结果截图')
            student_relative2_travel_image = student_info.get('同住人2行程码截图')
            student_relative3_name = student_info.get('同住人3的姓名')
            student_relative3_result_image = student_info.get('同住人3的24小时核酸检测结果截图')
            student_relative3_travel_image = student_info.get('同住人3行程码截图')
            student_relative4_name = student_info.get('同住人4的姓名')
            student_relative4_result_image = student_info.get('同住人4的24小时核酸检测结果截图')
            student_relative4_travel_image = student_info.get('同住人4行程码截图')
            student_relative5_name = student_info.get('同住人5的姓名')
            student_relative5_result_image = student_info.get('同住人5的24小时核酸检测结果截图')
            student_relative5_travel_image = student_info.get('同住人5行程码截图')

        # 识别图片
        total = ""
        file_name_and_date = get_excel_file_info(file_path)
        print("==============================================================")
        if student_name != '':
            name = student_name
            name_type = 0
            if notice_date == '20220605':
                # 第一张图
                img_path = student_result_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)
                # 第二张图
                # img_path = student_travel_image
                # total = do_ocr(img_path)
                # update_info(file_name_and_date,name, name_type, total)
            else:
                # 第一张图
                img_path = student_result_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

        if student_relative1_name != '':
            name = student_relative1_name
            name_type = 1
            if notice_date == '20220605':
                img_path = student_relative1_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)
            else:
                img_path = student_relative1_result_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

                img_path = student_relative1_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

        if student_relative2_name != '':
            name = student_relative2_name
            name_type = 2
            if notice_date == '20220605':
                img_path = student_relative2_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)
            else:
                img_path = student_relative2_result_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

                img_path = student_relative2_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

        if student_relative3_name != '':
            name = student_relative3_name
            name_type = 3
            if notice_date == '20220605':
                img_path = student_relative3_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)
            else:
                img_path = student_relative3_result_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

                img_path = student_relative3_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

        if student_relative4_name != '':
            name = student_relative4_name
            name_type = 4
            if notice_date == '20220605':
                img_path = student_relative4_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)
            else:
                img_path = student_relative4_result_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

                img_path = student_relative4_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

        if student_relative5_name != '':
            name = student_relative5_name
            name_type = 5
            if notice_date == '20220605':
                img_path = student_relative5_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)
            else:
                img_path = student_relative5_result_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

                img_path = student_relative5_travel_image
                total = do_ocr(img_path)
                update_info(file_name_and_date, name, name_type, total)

        df = df.append(info_dict, ignore_index=True)
        info_dict.clear()
    save_to_file(df, file_path)


def do_ocr(img_path):
    print("正在识别图像：", img_path)
    total = ''
    results = ocr.ocr(img_path, cls=False)
    for line in results:
        total += line[1][0] + " "
    print("原始识别出来的文字: ", total)
    # 处理下total中的特殊字符
    # 再有遇到特殊字符不好处理的可以在此添加

    total = total.replace(' o 0  ', '')
    total = total.replace('o', '')
    total = total.replace('○', '')
    total = total.replace('O', '')
    total = total.replace('③', '')
    total = total.replace('√', '')
    total = total.replace('已完成全程接种 ', '')
    total = total.replace('「', ' ')
    total = total.replace('「', '')
    total = total.replace('」', ' ')
    total = total.replace('丨', ' ')
    total = total.replace('》', '')
    total = total.replace('>', '')
    total = total.replace('<', '')
    total = re.sub(r'[A-Z][a-z]*', '', total)
    total = 'BEGIN ' + total  # 在开头加个BEGIN字样，后面有些特殊情况会使用
    print("初步处理后的文字：", total)
    return total


def update_info(file_name_and_date, name, name_type, total):
    # 时间范围需要从文件名读取
    start_date = file_name_and_date[1]
    last_date = file_name_and_date[2]

    validate = '不及格'
    city_total = ''
    contains_star = '否'
    image_type = ''
    result_ocr = ''
    name_ocr = ''
    final_result = ''
    sample_time_ocr = ''
    test_time_ocr = ''
    vaccine_name_ocr = ''
    vaccine_time_ocr = ''
    validate_travel = '不及格'
    if total.__contains__('粤康码信息'):
        image_type = 'CANTON'
        print("图片类型：", '我的粤康码信息')
        name_ocr = match(r'我的粤康码信息\s*(\S*)', total)
        sample_time_ocr = match(r'采样时间\s*(\S*)', total)
        test_time_ocr = match(r'检测时间\s*(\S*)', total)
        result_ocr = match(r'检测结果\s*(\S*)', total.replace(name_ocr, ''))
        if check_date(start_date, last_date, sample_time_ocr):
            validate = '时间及格'
        if result_ocr != '阴性':
            validate += "\n" + "检测结果没有‘阴性’字样，需要注意！"
        print("姓名: ", name_ocr)
        print("采样时间: ", sample_time_ocr)
        print("检测时间: ", test_time_ocr)
        print("检测结果: ", result_ocr)
        print("对比结果:", validate)

    elif total.__contains__('亲属出示') or total.__contains__('管理') or total.__contains__('播报'):
        image_type = 'QRCODE'
        print("图片类型：", '粤康码首页')
        name_ocr = match(r'深圳\s*(\S*)', total)
        result_ocr = match(r'新冠疫苗\s*(\S*)', total)
        test_time_ocr = match(r'阴性\s*(\S*)', total)  # 不是阴性我料你也不敢提交啊
        if result_ocr == '24':
            if int(last_date) == format_date(test_time_ocr):
                validate = '24小时'
            else:
                validate = '不是最后一天的检测时间，请注意！'
        elif result_ocr == '48小时':
            if int(last_date) == format_date(test_time_ocr):
                validate = '不是24小时内核酸结果，请注意！'
        else:
            validate = '不是48小时内核酸结果，请注意！'
        print("姓名: ", name_ocr)
        print("检测时间: ", test_time_ocr)
        print("检测结果: " + result_ocr)
        print("对比结果:", validate)

    elif total.__contains__('我的核酸检测记录'):
        image_type = 'MY_RECORD'
        print("图片类型：", '我的核酸检测记录')
        if total.count('我的核酸检测记录') > 1:  # 通常截图会有两个‘我的核酸检测记录'，也有的情况是只有一个
            if match(r'我的核酸检测记录\s*(\S*)', total) == '我的核酸检测记录':
                total = total.replace('我的核酸检测记录', ' ', 1)  # 只需要一次’我的核酸检测记录‘字串
                total = total.replace(match(r'我的核酸检测记录\s*(\S*)', total), ' ', 1)  # 去掉多余的干扰信息
            else:
                total = total.replace('我的核酸检测记录', ' ', 1)  # 只需要一次’我的核酸检测记录‘字串
                total = total.replace('刷新', ' ', 1)  # 去掉多余的干扰信息
        print("处理我的核酸检测记录的识别结果后的文字： " + total)
        only_chinese = re.sub(r'[^\u4e00-\u9fa5]+', ' ', total)
        print("只有中文的文字： " + only_chinese)
        name_ocr = match(r'我的核酸检测记录\s*(\S*)', only_chinese)
        sample_time_ocr = match(r'采样时间：\s*(\S*)', total)
        test_time_ocr = match(r'检测时间：\s*(\S*)', total)
        total = total.replace(name_ocr, ' ')  # 去掉名字，让’结果'字串排在'我的核酸检测记录'后面
        result_ocr = match(r'我的核酸检测记录\s*(\S*)', total)
        if result_ocr.__contains__('未出结果'):  # 未出结果的话，检测时间也会没有。
            test_time_ocr = ''  # 避免拿到前次的检测时间
        if check_date(start_date, last_date, sample_time_ocr):
            validate = '时间及格'
        if result_ocr != '阴性':
            validate += "\n" + "检测结果没有‘阴性’字样，需要注意！"
        if name_ocr != name:
            validate += "\n" + "名字不一致，需要注意。（可能传错图或者识别错误）"
        print("姓名: " + name_ocr)
        print("采样时间: " + sample_time_ocr)
        print("检测时间: " + test_time_ocr)
        print("检测结果: " + result_ocr)
        print("对比结果:", validate)

    elif total.__contains__('核酸检测记录'):
        image_type = 'RECORD'
        print("图片类型：", '核酸检测记录')
        if total.count('核酸检测记录') > 1:  # 通常截图会有两个‘核酸检测记录'，也有的情况是只有一个
            total = total.replace('核酸检测记录', ' ', 1)  # 只需要一次’核酸检测记录‘字串
        total = total.replace('检测中', ' ', 1)  # 去掉多余的干扰信息
        total = total.replace('检测完成', ' ', 1)  # 去掉多余的干扰信息
        total = total.replace('刷新', ' ', 1)  # 去掉多余的干扰信息
        print("处理核酸检测记录的识别结果后的文字： " + total)
        only_chinese = re.sub(r'[^\u4e00-\u9fa5]+', ' ', total)
        print("只有中文的文字： " + only_chinese)
        name_ocr = match(r'核酸检测记录\s*(\S*)', only_chinese)
        sample_time_ocr = match(r'采样时间\s*(\S*)', total)
        test_time_ocr = match(r'检测时间\s*(\S*)', total)
        result_ocr = match(r'检测结果\s*(\S*)', total)

        # 特殊处理，有的截图被音量键挡住了
        if result_ocr.__contains__('检测中'):
            if not sample_time_ocr:
                sample_time_ocr = match(r'间\s*(\S*)', total)

        if check_date(start_date, last_date, sample_time_ocr):
            validate = '时间及格'
        if result_ocr != '阴性':
            validate += "\n" + "检测结果没有‘阴性’字样，需要注意！"
        if name_ocr != name:
            validate += "\n" + "名字不一致，需要注意。（可能传错图或者识别错误）"
        print("姓名: " + name_ocr)
        print("采样时间: " + sample_time_ocr)
        print("检测时间: " + test_time_ocr)
        print("检测结果: " + result_ocr)
        print("对比结果:", validate)

    elif total.__contains__('通信大数据') or total.__contains__('绿色行程卡'):
        image_type = 'TRAVEL'
        print("图片类型：", '通信行程卡')
        total = total.replace('新于：', '新于： ')  # '新于：'加多个空格
        phone_ocr = match(r'请收下绿色行程卡\s*(\S*)', total).replace('的动态行程卡', '')
        update_time_ocr = match(r'新于：\s*(\S*)', total)
        if not update_time_ocr:
            update_time_ocr = match(r'更新：\s*(\S*)', total)
        if total.__contains__('市*'):
            contains_star = '是'
        total = total.replace(' ', '')  # 去掉空格
        city_orc = re.findall(r'省(.*?)市', total)  # 获得城市名称
        for city in city_orc:  # 列出城市
            city_total += city + ' '

        # 特殊处理：如果有可能，去掉多余的字
        city_total.replace('结果包含您在前14天内到访的国家（地区）与停留4小时以上的国内城', '')

        if check_date(start_date, last_date, update_time_ocr):
            if contains_star == '否':
                validate_travel = '时间及格'
        if len(city_orc) > 1:
            validate_travel = '包含深圳以外城市，请注意！'
        if '深' not in city_total:  # 地级市带“深”的也就深圳
            validate_travel = '没有包含深圳，请注意！'

        # 格式化一下时间
        update_time_ocr = update_time_ocr[:10] + ' ' + update_time_ocr[10:]
        final_result = phone_ocr + '\n' + update_time_ocr + '\n' + city_total + '\n是否带星: ' + contains_star + '\n' + validate_travel
        print("手机号： " + phone_ocr)
        print("更新时间：" + update_time_ocr)
        print("14天内到达或途径：", city_total)
        print("是否带*： " + contains_star)
        print("对比结果:", final_result)

    elif total.__contains__('疫苗接种记录'):
        image_type = 'VACCINE'
        print("图片类型：", '疫苗接种记录')
        if total.count('新冠疫苗接种记录') > 1:  # 通常截图会有两个‘我的核酸检测记录'，也有的情况是只有一个
            total = total.replace('新冠疫苗接种记录', ' ', 1)  # 只需要一次’我的核酸检测记录‘字串
        total = total.replace('刷新', ' ', 1)  # 去掉多余的干扰信息
        name_ocr = match(r'新冠疫苗接种记录\s*(\S*)', total)
        vaccine_name_ocr = match(r'疫苗名称\s*(\S*)', total)
        vaccine_time_ocr = match(r'接种时间\s*(\S*)', total)
        print("姓名: " + name_ocr)
        print("疫苗名称: " + vaccine_name_ocr)
        print("接种时间: " + vaccine_time_ocr)
    else:
        image_type = 'UNKNOWN'
        print("图片类型： ", '没有匹配已知的截图类型')
        print('尽量识别中....请注意该图片')
        total = total.replace('：', ' ')
        name_ocr = match(r'检测中\s*(\S*)', total)
        if not name_ocr:
            name_ocr = match(r'BEGIN\s*(\S*)', total)
        sample_time_ocr = match(r'采样时间\s*(\S*)', total)
        test_time_ocr = match(r'检测时间\s*(\S*)', total)
        result_ocr = match(r'检测结果\s*(\S*)', total)
        if not result_ocr:
            total = total.replace(name_ocr, ' ')  # 去掉多余的干扰信息
            result_ocr = match(r'BEGIN\s*(\S*)', total)
        if check_date(start_date, last_date, sample_time_ocr):
            validate = '时间及格'
        if result_ocr != '阴性':
            validate += "\n" + "检测结果没有‘阴性’字样，需要注意！"
        if sample_time_ocr == '' or test_time_ocr == '':
            validate = '无法判断时间，判为不及格，请注意！'
        print("姓名: " + name_ocr)
        print("采样时间: " + sample_time_ocr)
        print("检测时间: " + test_time_ocr)
        print("检测结果: " + result_ocr)
        print("对比结果:", validate)

    # 时间格式加个空格
    if sample_time_ocr != '':
        sample_time_ocr = sample_time_ocr[:10] + ' ' + sample_time_ocr[10:]
    if test_time_ocr != '':
        test_time_ocr = test_time_ocr[:10] + ' ' + test_time_ocr[10:]

    if name_type == 0:
        info_dict["学生姓名"] = name
        if image_type == 'TRAVEL':
            info_dict["学生行程码图片结果"] = final_result
        elif image_type == 'QRCODE':
            info_dict["学生核酸图片结果"] = "{0}\n检测时间： {1}\n检测结果： {2}\n是否及格： {3}\n 无法判断采样时间，需要注意！".format(name_ocr,
                                                                                                   test_time_ocr,
                                                                                                   result_ocr, validate)
        else:
            info_dict["学生核酸图片结果"] = "{0}\n采样时间： {1}\n检测时间： {2}\n检测结果： {3}\n是否及格： {4}".format(name_ocr, sample_time_ocr,
                                                                                             test_time_ocr, result_ocr,
                                                                                             validate)
    elif name_type == 1:
        info_dict["同住人1姓名"] = name
        if image_type == 'TRAVEL':
            info_dict["同住人1行程码图片结果"] = final_result
        elif image_type == 'VACCINE':
            info_dict["同住人1核酸图片结果"] = "这是疫苗接种证明，请查看备注。"
            info_dict["备注说明"] = "这是疫苗接种证明\n" + "姓名： " + name_ocr + "\n" + "疫苗名称: " \
                                + vaccine_name_ocr + "\n" + "接种时间: " + vaccine_time_ocr
        else:
            info_dict["同住人1核酸图片结果"] = "{0}\n采样时间： {1}\n检测时间： {2}\n检测结果： {3}\n是否及格： {4}".format(name_ocr,
                                                                                               sample_time_ocr,
                                                                                               test_time_ocr,
                                                                                               result_ocr,
                                                                                               validate)
    elif name_type == 2:
        info_dict["同住人2姓名"] = name
        if image_type == 'TRAVEL':
            info_dict["同住人2行程码图片结果"] = final_result
        elif image_type == 'VACCINE':
            info_dict["同住人2核酸图片结果"] = "这是疫苗接种证明，请查看备注。"
            info_dict["备注说明"] = "这是疫苗接种证明\n" + "姓名： " + name_ocr + "\n" + "疫苗名称: " \
                                + vaccine_name_ocr + "\n" + "接种时间: " + vaccine_time_ocr
        else:
            info_dict["同住人2核酸图片结果"] = "{0}\n采样时间： {1}\n检测时间： {2}\n检测结果： {3}\n是否及格： {4}".format(name_ocr,
                                                                                               sample_time_ocr,
                                                                                               test_time_ocr,
                                                                                               result_ocr,
                                                                                               validate)
    elif name_type == 3:
        info_dict["同住人3姓名"] = name
        if image_type == 'TRAVEL':
            info_dict["同住人3行程码图片结果"] = final_result
        elif image_type == 'VACCINE':
            info_dict["同住人3核酸图片结果"] = "这是疫苗接种证明，请查看备注。"
            info_dict["备注说明"] = "这是疫苗接种证明\n" + "姓名： " + name_ocr + "\n" + "疫苗名称: " \
                                + vaccine_name_ocr + "\n" + "接种时间: " + vaccine_time_ocr
        else:
            info_dict["同住人3核酸图片结果"] = "{0}\n采样时间： {1}\n检测时间： {2}\n检测结果： {3}\n是否及格： {4}".format(name_ocr,
                                                                                               sample_time_ocr,
                                                                                               test_time_ocr,
                                                                                               result_ocr,
                                                                                               validate)
    elif name_type == 4:
        info_dict["同住人4姓名"] = name
        if image_type == 'TRAVEL':
            info_dict["同住人4行程码图片结果"] = final_result
        elif image_type == 'VACCINE':
            info_dict["同住人4核酸图片结果"] = "这是疫苗接种证明，请查看备注。"
            info_dict["备注说明"] = "这是疫苗接种证明\n" + "姓名： " + name_ocr + "\n" + "疫苗名称: " \
                                + vaccine_name_ocr + "\n" + "接种时间: " + vaccine_time_ocr
        else:
            info_dict["同住人4核酸图片结果"] = "{0}\n采样时间： {1}\n检测时间： {2}\n检测结果： {3}\n是否及格： {4}".format(name_ocr,
                                                                                               sample_time_ocr,
                                                                                               test_time_ocr,
                                                                                               result_ocr,
                                                                                               validate)
    elif name_type == 5:
        info_dict["同住人5姓名"] = name
        if image_type == 'TRAVEL':
            info_dict["同住人5行程码图片结果"] = final_result
        elif image_type == 'VACCINE':
            info_dict["同住人5核酸图片结果"] = "这是疫苗接种证明，请查看备注。"
            info_dict["备注说明"] = "这是疫苗接种证明\n" + "姓名： " + name_ocr + "\n" + "疫苗名称: " \
                                + vaccine_name_ocr + "\n" + "接种时间: " + vaccine_time_ocr
        else:
            info_dict["同住人5核酸图片结果"] = "{0}\n采样时间： {1}\n检测时间： {2}\n检测结果： {3}\n是否及格： {4}".format(name_ocr,
                                                                                               sample_time_ocr,
                                                                                               test_time_ocr,
                                                                                               result_ocr,
                                                                                               validate)


# 文件名必须是两个日期来命名：xxx-20220407-20220410.xlsx
# 返回最新创建时间的文件
def get_excel_file_info(file_path):
    file_name_and_date = []
    start_date = file_path[-22:-14]
    last_date = file_path[-13:-5]
    file_name_and_date.append(file_path)  # 文件名
    file_name_and_date.append(start_date)  # 开始日期
    file_name_and_date.append(last_date)  # 结束日期
    print("表格名称参数：", file_name_and_date)
    return file_name_and_date


if __name__ == '__main__':
    # f = open('LOG-20220410.txt', 'a')
    # sys.stdout = f
    # sys.stderr = f
    start_time = time.time()
    fire.Fire(deal_file)
    end_time = time.time()
    print("运行时间:", end_time - start_time)
    # f.close()
